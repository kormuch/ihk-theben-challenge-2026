"""
Extract raw text from uploaded files for LLM processing.
Supports: PDF, CSV, JSON, XML, XLSX, TXT/MD, images (PNG/JPG/etc via OCR).
"""
import csv
import json
import io
from pathlib import Path

IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}


def extract_text(file_path: Path) -> str:
    """Extract text content from a file. Returns plain text string."""
    suffix = file_path.suffix.lower()

    if suffix == ".pdf":
        return _extract_pdf(file_path)
    elif suffix in IMAGE_SUFFIXES:
        return _extract_image(file_path)
    elif suffix in (".csv", ".tsv"):
        return _extract_csv(file_path)
    elif suffix == ".json":
        return _extract_json(file_path)
    elif suffix == ".xml":
        return file_path.read_text(encoding="utf-8", errors="replace")
    elif suffix in (".xlsx", ".xls"):
        return _extract_xlsx(file_path)
    elif suffix in (".txt", ".md", ".log"):
        return file_path.read_text(encoding="utf-8", errors="replace")
    else:
        # Try reading as text
        try:
            return file_path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            raise ValueError(f"Cannot extract text from {suffix} files")


def _ocr_image(file_path: Path) -> str:
    """Run OCR on an image file. Returns extracted text."""
    import pytesseract
    from PIL import Image

    img = Image.open(file_path)
    text = pytesseract.image_to_string(img, lang="deu+eng")
    return text.strip()


def _extract_image(file_path: Path) -> str:
    """Extract text from an image via OCR."""
    text = _ocr_image(file_path)
    if not text:
        raise ValueError(f"OCR could not extract text from {file_path.name}")
    return f"[OCR: {file_path.name}]\n{text}"


def _extract_pdf(file_path: Path) -> str:
    import pdfplumber

    parts = []
    with pdfplumber.open(file_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if text and text.strip():
                parts.append(f"[Page {i}]\n{text.strip()}")
            tables = page.extract_tables()
            for table in tables:
                if table and len(table) >= 2:
                    rows = []
                    for row in table:
                        cells = [str(c).strip() if c else "" for c in row]
                        rows.append(" | ".join(cells))
                    parts.append(f"[Table, Page {i}]\n" + "\n".join(rows))

    # OCR fallback: if pdfplumber found no text, try OCR on each page
    if not parts:
        parts = _ocr_pdf_fallback(file_path)

    return "\n\n".join(parts)


def _ocr_pdf_fallback(file_path: Path) -> list[str]:
    """Convert PDF pages to images and run OCR. Used for scanned PDFs."""
    import pytesseract
    from PIL import Image

    parts = []
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(file_path)
    except ImportError:
        # pdf2image not available — try pdfplumber page-to-image conversion
        import pdfplumber
        images = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                img = page.to_image(resolution=300).original
                images.append(img)

    for i, img in enumerate(images, start=1):
        text = pytesseract.image_to_string(img, lang="deu+eng")
        if text and text.strip():
            parts.append(f"[Page {i}, OCR]\n{text.strip()}")

    return parts


def _extract_csv(file_path: Path) -> str:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append(" | ".join(row))
    return "\n".join(rows)


def _extract_json(file_path: Path) -> str:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    try:
        data = json.loads(text)
        return json.dumps(data, indent=2, ensure_ascii=False)
    except json.JSONDecodeError:
        return text


def _extract_xlsx(file_path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)
